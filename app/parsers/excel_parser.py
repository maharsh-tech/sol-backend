from openpyxl import load_workbook


def parse_excel(file_path: str, document_name: str) -> list[dict]:
    """Parse an Excel file and return a list of text chunks with metadata."""
    results: list[dict] = []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cell_values = [str(cell) if cell is not None else "" for cell in row]
            line = " | ".join(cell_values).strip()
            if line and line != "|".join([""] * len(cell_values)):
                rows_text.append(line)
        text = "\n".join(rows_text).strip()
        if text:
            results.append(
                {
                    "text": text,
                    "metadata": {
                        "document_name": document_name,
                        "page": sheet_name,
                        "source_type": "excel",
                    },
                }
            )
    wb.close()
    return results
